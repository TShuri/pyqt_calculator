import pandas as pd
from openpyxl import load_workbook

class Logic():
    def __init__(self, output_func=None, ask_gp_callback=None):
        self.name_obyz = [] # Название обязательства
        self.su = []
        self.od = [] # Основной долг
        self.penod = []
        self.prc = []
        self.penprc = []
        self.gp = []
        self.km = []
        self.pp = []
        self.prsud = []
        self.namesdolg = []
        self.neus = []
        self.temp = []
        self.namesdolgsumm = []
        
        self.output = output_func if output_func else self.output
        self.ask_gp_callback = ask_gp_callback

    # --- Сброс всех списков ---
    def clear_all(self, event=None):
        self.su.clear()
        self.od.clear()
        self.prc.clear()
        self.gp.clear()
        self.km.clear()
        self.pp.clear()
        self.penod.clear()
        self.penprc.clear()

    def clear_list_and_frame(self):
        self.neus.clear()
        self.name_obyz.clear()
        self.su.clear()
        self.od.clear()
        self.prc.clear()
        self.gp.clear()
        self.km.clear()
        self.pp.clear()
        self.penod.clear()
        self.penprc.clear()

    # --- Проверка госпошлин ---
    def proverkagpkd(self, wb): # Проверка госпошлины кредитного договора
        for i in (wb.sheetnames):
            if i == "Расчёт":
                sheet = wb["Расчёт"]
                max_rows = sheet.max_row
                for i in range(17, max_rows):
                    sku = sheet.cell(row=6, column=9).value
                    try:
                        if sku > 0:
                            self.output(f'{sheet["a1"].value}, НАЙДЕНЫ НАЧИСЛЕНИЯ ПО ГОСПОШЛИНЕ! ОТКРОЙТЕ Расчет')
                            break
                    except TypeError:
                        pass

    def proverkagpkk(self, wb): # Проверка госпошлины кредитной карты
        for i in (wb.sheetnames):
            if i == "Приложение 5":
                sheet = wb["Приложение 5"]
                max_rows = sheet.max_row
                for i in range(17, max_rows):
                    sku = sheet.cell(row=11, column=5).value
                    try:
                        if sku > 0:
                            self.output(f'{sheet["a7"].value}, "НАЙДЕНЫ НАЧИСЛЕНИЯ ПО ГОСПОШЛИНЕ! ОТКРОЙТЕ Расчет_7')
                            break
                    except TypeError:
                        break

    def proverkagp(self, wb): # Проверка госпошлины
        for i in (wb.sheetnames):
            if i == "Расчет_7":
                sheet = wb["Расчет_7"]
                max_rows = sheet.max_row
                for i in range(17, max_rows):
                    sku = sheet.cell(row=17, column=10).value
                    try:
                        if sku > 0:
                            self.output(f'{sheet["a7"].value}, "НАЙДЕНЫ НАЧИСЛЕНИЯ ПО ГОСПОШЛИНЕ! ОТКРОЙТЕ Расчет_7')
                            break
                    except TypeError:
                        break

    # --- Преобразование строковых значений в float ---
    def to_float(sums: list) -> None:
        sums[:] = [float(x) for x in sums if str(x).strip() != "–"]

    # --- Вопрос пользователю: учитывать ли госпошлину ---
    def ask_into_gp(self, msg, val): # Вопрос пользователю: учитывать ли госпошлину
        gp_msg = f'Учесть госпошлину при расчетах \n{msg} \nв сумме {val:.2f}?'
        if self.ask_gp_callback:
            return 'yes' if self.ask_gp_callback(gp_msg) else 'no'
        else:
            return 'no'

    # --- Основная функция для расчета кредитной карты ---
    def run(self, files):
        for file in files:
            try:
                self.output(f'\nОбязательство {file}')
                try:
                    wb = load_workbook(file, data_only=True)
                except:
                    df = pd.read_excel(file)
                    df.to_excel(file, index=False, header=False)
                    wb = load_workbook(file, data_only=True)
                for sheet in (wb.sheetnames):
                    if sheet == "Титульный лист" or sheet == "Sheet1":
                        try:
                            wb.active = wb["Титульный лист"]
                            ws = wb.active
                        except:
                            wb.active = wb["Sheet1"]
                            ws = wb.active
                        if type(ws["a5"].value) == str:
                            self.namesdolg.append(ws["a5"].value[20:])
                        else:
                            self.namesdolg.append(ws["a2"].value[20:])

                        # Проверка госпошлин
                        self.proverkagp(wb)
                        self.proverkagpkd(wb)
                        self.proverkagpkk(wb)

                        for row in range(3, 27):
                            if (ws[row][0].value == "Задолженность по основному долгу (ссудная задолженность)"
                                    or ws[row][0].value == "Ссудная задолженность"
                                    or ws[row][0].value == "присужденный основной долг"
                                    or ws[row][0].value == "Задолженность по кредиту"
                                    or ws[row][0].value == "Просроченный основной долг, списанный за счет резерва"):

                                if len(ws[row]) >= 9 and ws[row][8] != None:
                                    val = ws[row][8].value
                                    if type(val) == str:
                                        val = float(val.replace(",", "."))
                                    if val > 0:  # не убирай if
                                        if (val in self.temp):
                                            continue
                                        else:
                                            self.od.append(val)
                                            self.su.append(val)
                                            self.temp.append(val)
                                elif len(ws[row]) >= 9 and ws[row][7].value != None:
                                    val = ws[row][8].value
                                    if type(val) == str:
                                        val = float(val.replace(",", "."))
                                    if val in self.od:
                                        continue
                                    else:
                                        self.od.append(val)
                                        self.su.append(val)
                                        self.temp.append(val)
                                        continue
                                else:
                                    val = ws[row][5].value
                                    if type(val) == str:
                                        val = float(val.replace(",", "."))
                                    if val > 0:
                                        if (val in self.temp):
                                            continue
                                        else:
                                            self.od.append(val)
                                            self.su.append(val)
                                            self.temp.append(val)

                            if (ws[row][0].value == "Задолженность по процентам"
                                    or ws[row][0].value == "Проценты за кредит"
                                    or ws[row][0].value == "присужденные просроченные проценты на просроченный основной долг"
                                    or ws[row][0].value == "присужденные просроченные проценты"):
                                if len(ws[row]) >= 9 and ws[row][8].value != None:
                                    val = ws[row][8].value
                                    if type(val) == str:
                                        val = float(val.replace(",", "."))
                                    if val > 0:  # не убирай if
                                        self.prc.append(val)
                                        self.su.append(val)
                                        self.temp.append(val)
                                elif len(ws[row]) >= 8 and ws[row][7].value != None:
                                    val = ws[row][7].value
                                    if type(val) == str:
                                        val = float(val.replace(",", "."))
                                    if val > 0:  # не убирай if
                                        self.prc.append(val)
                                        self.su.append(val)
                                        self.temp.append(val)
                                else:  # не убирай if
                                    val = ws[row][5].value
                                    if type(val) == str:
                                        val = float(val.replace(",", "."))
                                    if val > 0:
                                        self.prc.append(val)
                                        self.su.append(val)
                                        self.temp.append(val)
                            if (ws[row][0].value == "Сумма госпошлин, списанных за счет резерва"
                                    or ws[row][0].value == "Госпошлина"):
                                if len(ws[row]) >= 9 and ws[row][8].value != None:
                                    val = ws[row][8].value
                                    if type(val) == str:
                                        val = float(val.replace(",", "."))
                                    if (val > 0):  # не убирай if
                                        msg = f'{ws["a5"].value[20:]}'
                                        self.output(msg)
                                        self.output('Учесть госпошлину при расчетах?')
                                        res = self.ask_into_gp(msg=msg, val=val)
                                        if (res == "yes"):
                                            if val > 0:  # не убирай if
                                                self.gp.append(val)
                                                self.temp.append(val)
                                                self.su.append(val)
                                                self.output(f'Госпошлина в сумме {val} посчитана')
                                                self.output("")
                                                continue
                                        else:
                                            if val > 0:  # не убирай if
                                                self.output(f'Госпошлина в сумме {val} не посчитана')
                                                self.output("")
                                                continue

                                if len(ws[row]) >= 9 and ws[row][7].value != None:
                                    val = ws[row][7].value
                                    if type(val) == str:
                                        val = float(val.replace(",", "."))
                                    if (val > 0):  # не убирай if
                                        msg = f'{ws["a5"].value[20:]}'
                                        self.output(msg)
                                        self.output('Учесть госпошлину при расчетах?')
                                        res = self.ask_into_gp(msg=msg, val=val)
                                        if (res == "yes"):
                                            if val > 0:  # не убирай if
                                                self.gp.append(val)
                                                self.temp.append(val)
                                                self.su.append(val)
                                                self.output(f'Госпошлина в сумме {val} посчитана')
                                                self.output("")
                                                continue
                                        else:
                                            if val > 0:  # не убирай if
                                                self.output(f'Госпошлина в сумме {val} не посчитана')
                                                self.output("")
                                                continue

                                if len(ws[row]) >= 6:
                                    val = ws[row][5].value
                                    if type(val) == str:
                                        val = float(val.replace(",", "."))
                                    if (val > 0):  # не убирай if
                                        msg = f'{ws["a2"].value[20:]}'
                                        self.output(msg)
                                        self.output('Учесть госпошлину при расчетах?')
                                        res = self.ask_into_gp(msg=msg, val=val)
                                        if (res == "yes"):
                                            if val > 0:  # не убирай if
                                                self.gp.append(val)
                                                self.temp.append(val)
                                                self.su.append(val)
                                                self.output(f'Госпошлина в сумме {val} посчитана')
                                                self.output("")
                                                continue
                                        else:
                                            if val > 0:  # не убирай if
                                                self.output(f'Госпошлина в сумме {val} не посчитана')
                                                self.output("")
                                                continue

                            if (ws[row][0].value == "Комиссии на отчетную дату"
                                    or ws[row][0].value == "Комиссия за пользование картой"
                                    or ws[row][0].value == "Сумма списанных коммиссий за счет резерва"):
                                self.km.append(ws[row][8].value)
                                self.temp.append(ws[row][8].value)
                                self.su.append(ws[row][8].value)
                                continue
                            if (ws[row][0].value == "Неустойка по кредиту" or ws[row][0].value == "Неустойка по процентам"
                                    or ws[row][0].value == "Неустойка за несвоевременное погашение Обязательного платежа"
                                    or ws[row][0].value == "Неустойки, признанные должником в дату реструктуризации/мирового соглашения по банковской карте"
                                    or ws[row][0].value == "Неустойки (присужденные)"
                                    or ws[row][0].value == "Сумма неустоек, списанных за счет резерва"
                                    or ws[row][0].value == "Неустойки за неисполнение условий договора"
                                    or ws[row][0].value == "Списанные неустойки"
                                    or ws[row][0].value == "присужденные неустойки по процентам"
                                    or ws[row][0].value == "присужденные неустойки по кредиту"
                                    or ws[row][0].value == "Сумма неустоек за просроченный основной долг, списанных за счет резерва"
                                    or ws[row][0].value == "Неустойка за просроченные проценты"):
                                if len(ws[row]) >= 9 and ws[row][8].value != None:
                                    val = ws[row][8].value
                                    if type(val) == str:
                                        val = float(val.replace(",", "."))
                                    if (val > 0):
                                        self.neus.append(val)
                                        self.temp.append(val)
                                        self.su.append(val)
                                        continue

                                if len(ws[row]) >= 9 and ws[row][7].value != None:
                                    val = ws[row][7].value
                                    if type(val) == str:
                                        val = float(val.replace(",", "."))
                                    if (val > 0):
                                        self.neus.append(val)
                                        self.temp.append(val)
                                        self.su.append(val)
                                        continue

                                if len(ws[row]) >= 6:
                                    val = ws[row][5].value
                                    if type(val) == str:
                                        val = float(val.replace(",", "."))
                                    if (val > 0):
                                        self.neus.append(val)
                                        self.temp.append(val)
                                        self.su.append(val)
                                        continue
                            try:
                                if ("по госпошлине" in ws[row][0].value):
                                    msg = f'{ws["a5"].value[20:]}'
                                    self.output(msg)
                                    self.output('Учесть госпошлину при расчетах?')
                                    res = self.ask_into_gp(msg=msg, val=val)
                                    if (res == "yes"):
                                        self.gp.append(ws[row][8].value)
                                        self.temp.append(ws[row][8].value)
                                        self.su.append(ws[row][8].value)
                                        self.output(f'Госпошлина в сумме {ws[row][8].value} посчитана')
                                        self.output("")
                                        continue
                                    else:
                                        self.output(f'Госпошлина в сумме {ws[row][8].value} не посчитана')
                                        self.output("")
                                        continue
                                if "Проценты за кредит " in ws[row][0].value and ws[row][7].value != None:
                                    if ws[row][7].value in self.prc:
                                        continue
                                    else:
                                        self.prc.append(ws[row][7].value)
                                        self.temp.append(ws[row][7].value)
                                        self.su.append(ws[row][7].value)
                                        continue
                                if "Ссудная задолженность " in ws[row][0].value and ws[row][7].value != None:
                                    if ws[row][7].value in self.od:
                                        continue
                                    else:
                                        self.od.append(ws[row][7].value)
                                        self.temp.append(ws[row][7].value)
                                        self.su.append(ws[row][7].value)
                                        continue
                            except:
                                continue
                        y = sum(self.temp)
                        if (ws["a1"].value == "Управление администрирования кредитов ЦСКО"):
                            self.name_obyz.append(ws["a5"].value[65:81])
                            self.output(f'{ws["a5"].value}')
                        elif (ws["a1"].value == "Управление администрирования кредитов ПЦП МСЦ"
                            or ws["a1"].value == "Управление администрирования кредитов МСЦ"):
                            self.name_obyz.append(ws["a5"].value[110:140])
                            self.output(f'{ws["a5"].value}')
                        elif (ws["g1"].value == "Подразделение по работе с проблемной задолженностью физических лиц"):
                            self.name_obyz.append(ws["a4"].value[110:140])
                            self.output(f'{ws["a4"].value}')
                        else:
                            if ws["a1"].value != None:
                                self.name_obyz.append(ws["a1"].value[186:217])
                                self.output(f'{ws["a1"].value}')
                            else:
                                self.name_obyz.append(ws["a2"].value[:201])
                                self.output(f'{ws["a2"].value[:201]}')

                        self.output(f'\nИтог - {y:.2f}')
                        self.output("")
                        self.namesdolgsumm.append(y)
                        self.temp.clear()

                    elif sheet == "Задолженность по договору":
                        ws = wb["Задолженность по договору"]
                        for row in range(3, 27):
                            if (ws[row][0].value == "Просроченная ссудная задолженность (присужденная)"
                                    or ws[row][0].value == "Просроченная ссудная задолженность"
                                    or ws[row][0].value == "Ссудная задолженность"
                                    or ws[row][0].value == "Основной долг на в/б, списанный за счет резерва"):
                                try:
                                    if ws[row][1].value > 0:  # не убирай if
                                        self.od.append(ws[row][1].value)
                                        self.su.append(ws[row][1].value)
                                        self.temp.append(ws[row][1].value)
                                except TypeError:
                                    self.od.append(ws[row][7].value)
                                    self.su.append(ws[row][7].value)
                                    self.temp.append(ws[row][7].value)
                                    continue
                            if (ws[row][0].value == "Задолженность по процентам"
                                    or ws[row][0].value == "Просроченная задолженность по процентам (присужденная)"
                                    or ws[row][0].value == "Неполученные списанные на в/б проценты"
                                    or ws[row][0].value == "Просроченная задолженность по процентам"):
                                try:
                                    if ws[row][1].value > 0:  # не убирай if
                                        self.prc.append(ws[row][1].value)
                                        self.su.append(ws[row][1].value)
                                        self.temp.append(ws[row][1].value)
                                except TypeError:
                                    self.prc.append(ws[row][7].value)
                                    self.su.append(ws[row][7].value)
                                    self.temp.append(ws[row][7].value)
                                    continue
                            if (ws[row][0].value == "Пени за проценты (присужденные)"):
                                try:
                                    if ws[row][1].value > 0:  # не убирай if
                                        self.penprc.append(ws[row][1].value)
                                        self.su.append(ws[row][1].value)
                                        self.temp.append(ws[row][1].value)
                                except TypeError:
                                    self.penprc.append(ws[row][7].value)
                                    self.su.append(ws[row][7].value)
                                    self.temp.append(ws[row][7].value)
                                    continue
                            if (ws[row][0].value == "Пени за кредит (присужденные)"):
                                try:
                                    if ws[row][1].value > 0:  # не убирай if
                                        self.penod.append(ws[row][1].value)
                                        self.su.append(ws[row][1].value)
                                        self.temp.append(ws[row][1].value)
                                except TypeError:
                                    self.penod.append(ws[row][7].value)
                                    self.su.append(ws[row][7].value)
                                    self.temp.append(ws[row][7].value)
                                continue
                            if (ws[row][0].value == "Неустойка по кредиту" or ws[row][0].value == "Неустойка по процентам"):
                                self.neus.append(ws[row][8].value)
                                self.temp.append(ws[row][8].value)
                                self.su.append(ws[row][8].value)
                                continue
                            if (ws[row][0].value == "Госпошлина"
                                    or ws[row][0].value == "Госпошлина (присужденная)"
                                    or ws[row][0].value == "Расходы на оплату третейского сбора"
                                    or ws[row][0].value == "Списанная на в/б госпошлина (присуждённая)"):
                                if (ws[row][8].value > 0):
                                    msg = f'{ws["a5"].value[20:]}'
                                    self.output(msg)
                                    self.output('Учесть госпошлину при расчетах?')
                                    res = self.ask_into_gp(msg=msg, val=ws[row][1].value)
                                    
                                    try:
                                        if (res == "yes"):
                                            self.gp.append(ws[row][1].value)
                                            self.temp.append(ws[row][1].value)
                                            self.su.append(ws[row][1].value)
                                            self.output(f'Госпошлина в сумме {ws[row][1].value} посчитана')
                                            self.output("")
                                        else:
                                            self.output(f'Госпошлина в сумме {ws[row][1].value} не посчитана')
                                            self.output("")
                                            continue
                                    except TypeError:
                                        if (ws[row][7].value > 0):
                                            msg = f'{ws["a5"].value[20:]}'
                                            self.output(msg)
                                            self.output('Учесть госпошлину при расчетах?')
                                            res = self.ask_into_gp(msg=msg, val=ws[row][7].value)
                                            
                                            if (res == "yes"):
                                                self.gp.append(ws[row][7].value)
                                                self.temp.append(ws[row][7].value)
                                                self.su.append(ws[row][7].value)
                                                self.output(f'Госпошлина в сумме {ws[row][7].value} посчитана')
                                                self.output("") 
                                            else:
                                                self.output(f'Госпошлина в сумме {ws[row][7].value} не посчитана')
                                                self.output("")
                                                
                        y = sum(set(self.temp))
                        
                        self.name_obyz.append(ws["c4"].value)

                        self.output(f'{'-' * 60}')
                        self.output(ws["a5"].value)
                        self.output(ws["c5"].value)
                        self.output(ws["c4"].value)

                        self.output(f'\nИтог - {y:.2f}')
                        
                        self.namesdolgsumm.append(y)
                        self.temp.clear()

                    elif sheet == "Лист1" or "Отчет по операциям" in sheet:
                        try:
                            ws = wb[sheet]
                            self.namesdolg.append(ws["a1"].value)
                        except KeyError:
                            self.output("Ошибка, Неверный РЦИ")
                            break
                        for row in range(17, 27):
                            if (ws[row][0].value == "Основной долг"):
                                try:
                                    x = ws[row][2].value.replace(" ", "")
                                    x = x.replace(",", ".")
                                    self.od.append(x)
                                    self.su.append(x)
                                    self.temp.append(x)
                                except TypeError:
                                    self.od.append(ws[row][2].value)
                                    self.su.append(ws[row][2].value)
                                    self.temp.append(ws[row][2].value)
                                    continue
                            if (ws[row][0].value == "Проценты за пользование кредитом"):
                                try:
                                    x = ws[row][2].value.replace(" ", "")
                                    x = x.replace(",", ".")
                                    self.prc.append(x)
                                    self.su.append(x)
                                    self.temp.append(x)
                                except:
                                    self.prc.append(ws[row][2].value)
                                    self.su.append(ws[row][2].value)
                                    self.temp.append(ws[row][2].value)
                                    continue
                            if (ws[row][0].value == "Неустойка за просроченную ссуду"):
                                x = ws[row][2].value.replace(" ", "")
                                x = x.replace(",", ".")
                                self.neus.append(x)
                                self.su.append(x)
                                self.temp.append(x)

                            if (ws[row][0].value == "Неустойка за просроченные проценты"):
                                x = ws[row][2].value.replace(" ", "")
                                x = x.replace(",", ".")
                                self.neus.append(x)
                                self.su.append(x)
                                self.temp.append(x)

                            if (ws[row][0].value == "Просроченные платежи"):
                                try:
                                    x = ws[row][2].value.replace(" ", "")
                                    x = x.replace(",", ".")

                                    self.pp.append(x)
                                    self.su.append(x)
                                    self.temp.append(x)
                                except:
                                    self.pp.append(ws[row][2].value)
                                    self.su.append(ws[row][2].value)
                                    self.temp.append(ws[row][2].value)
                                    continue
                        self.to_float(self.neus)
                        self.to_float(self.su)
                        self.to_float(self.temp)
                        self.to_float(self.od)
                        self.to_float(self.prc)
                        self.to_float(self.pp)
                        y = sum(self.temp)
                    
                        if (ws["a1"].value == "Управление администрирования кредитов ЦСКО"):
                            self.name_obyz.append(ws["a5"].value[65:81])
                            self.output(ws["a5"].value)
                            
                        elif (ws["a1"].value == "Управление администрирования кредитов ПЦП МСЦ"
                            or ws["a1"].value == "Управление администрирования кредитов МСЦ"):
                            self.name_obyz.append(ws["a5"].value[110:140])
                            self.output(ws["a5"].value)
                            
                        else:
                            self.name_obyz.append(ws["a1"].value[186:217])
                            self.output(ws["a1"].value)
                        
                        self.output(f'\nИтог - {y:.2f}')
                        self.output("")
                        # text.insert("1.0", "\n")
                        self.namesdolgsumm.append(y)
                        self.temp.clear()
                    break
            except Exception as e:
                self.output(f"Ошибка при обработке файла {file}")
                continue

        rcy = sum(self.su)
        pod = sum(self.od)
        pprc = sum(self.prc)
        pgp = sum(self.gp)
        kmp = sum(self.km)
        ppp = sum(self.pp)
        ppenprc = sum(self.penprc)
        ppenod = sum(self.penod)
        prsudp = sum(self.prsud)
        pneus = sum(self.neus)
        
        # self.output(f"{'-' * 15} Разбивка {'-' * 15}")
        # self.output(f"{'-' * 40}")
        
        total = {}
        
        if pneus > 0:
            total['Неустойки'] = pneus
            # self.output(f'{pneus:.2f} - неустойки')
        if ppenod > 0:
            total['Пени за кредит'] = ppenod
            # self.output(f'{ppenod:.2f} - пени за кредит')
        if ppenprc > 0:
            total['Пени за проценты'] = ppenprc
            # self.output(f'{ppenprc:.2f} - пени за проценты')
        if prsudp > 0:
            total['Прочие судебные расходы'] = prsudp
            # self.output(f'{prsudp:.2f} - прочие судебные расходы')
        if ppp > 0:
            total['Просроченная ссудная задолженность'] = ppp
            # self.output(f'{ppp:.2f} - просроченная ссудная задолженность')
        try:
            if kmp > 0:
                total['Комиссия'] = kmp
                # self.output(f'{kmp:.2f} - комиссия')
        except NameError:
            pass
        if pgp > 0:
            total['Госпошлина'] = pgp
            # self.output(f'{pgp:.2f} - госпошлина')
        if pprc > 0:
            total['Просроченные проценты'] = pprc
            # self.output(f'{pprc} - просроченные проценты')
        if pod > 0:
            total['Просроченный основной долг'] = pod
            # self.output(f'{pod} - просроченный основной долг')
        if rcy > 0:
            self.output(f"{'-' * 40}")
            total['Общая сумма'] = rcy
            # self.output(f'{rcy:.2f} - ОБЩАЯ СУММА')
            
        while len(self.name_obyz) > 1:
            self.name_obyz.pop()
            
        # for i in self.name_obyz:
        #     if i not in self.data["Обязательства"]:
        #         data["Обязательства"].append(i)
        #         data["Количество обязательств"].append(kol_obyz)
        #         # data["Время"].append(datetime.now())
        
        # self.output("")
        if (rcy < 100000):
            gosposhlina = 10000 / 2
            total['Оплата госпошлины'] = gosposhlina
            # self.output(f'{gosposhlina:.0f} - ОПЛАТА ГОСПОШЛИНЫ')
        elif (rcy > 100000 and rcy < 1000000):
            gosposhlina = ((rcy - 100000) * 0.05 + 10000) / 2
            total['Оплата госпошлины'] = gosposhlina
            # self.output(f'{gosposhlina:.0f} - ОПЛАТА ГОСПОШЛИНЫ')
        elif (rcy > 1000000 and rcy < 10000000):
            gosposhlina = ((rcy - 1000000) * 0.03 + 55000) / 2
            total['Оплата госпошлины'] = gosposhlina
            # self.output(f'{gosposhlina:.0f} - ОПЛАТА ГОСПОШЛИНЫ')
        elif (rcy > 10000000 and rcy < 50000000):
            gosposhlina = ((rcy - 10000000) * 0.01 + 325000) / 2
            total['Оплата госпошлины'] = gosposhlina
            # self.output(f'{gosposhlina:.0f} - ОПЛАТА ГОСПОШЛИНЫ')
        elif (rcy > 50000000):
            gosposhlina = ((rcy - 50000000) * 0.005 + 725000) / 2
            if (gosposhlina > 10000000):
                gosposhlina = 10000000
            total['Оплата госпошлины'] = gosposhlina
        #     self.output(f'{gosposhlina:.0f} - ОПЛАТА ГОСПОШЛИНЫ')
        # self.output(f"{'-' * 40}")    
        try:
            wb.close()
        except UnboundLocalError:
            self.output('Файлы не выбраны')
        
        return total

if __name__ == "__main__":
    logic = Logic()
    logic.run()
